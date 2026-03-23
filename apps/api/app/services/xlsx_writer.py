"""XLSX workbook writer (EXP-02, EXP-03, EXP-04)."""

import json
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl


def create_workbook() -> openpyxl.Workbook:
    """Create new workbook."""
    return openpyxl.Workbook()


def write_answers_to_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, answers: list[dict]) -> None:
    """Write answers to sheet. answers: [{row, col, text}, ...]."""
    for a in answers:
        row, col = a.get("row", 1), a.get("col", 1)
        ws.cell(row=row, column=col, value=a.get("text", ""))


def save_workbook(wb: openpyxl.Workbook) -> bytes:
    """Save workbook to bytes."""
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_workbook_from_bytes(data: bytes) -> openpyxl.Workbook:
    """Load workbook from bytes (EXP-03)."""
    return openpyxl.load_workbook(BytesIO(data), data_only=False)


def _copy_cell_style(source: openpyxl.cell.cell.Cell, target: openpyxl.cell.cell.Cell) -> None:
    """Copy formatting from source cell to target cell (EXP-04)."""
    if source.has_style:
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.number_format = copy(source.number_format)
        target.alignment = copy(source.alignment)


def place_answers_into_workbook(
    wb: openpyxl.Workbook,
    placements: list[dict[str, Any]],
    answer_col_offset: int = 1,
) -> openpyxl.Workbook:
    """Place answers into workbook cells (EXP-03, EXP-04).

    placements: list of {sheet, row, col, text}.
    row, col are 1-based (question cell from source_location).
    Answer is written to (row, col + answer_col_offset) by default (next column).
    Preserves formatting by copying style from the question cell to the answer cell.
    """
    for p in placements:
        sheet_name = p.get("sheet")
        row = p.get("row", 1)
        col = p.get("col", 1)
        text = p.get("text", "")
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        answer_col = col + answer_col_offset
        question_cell = ws.cell(row=row, column=col)
        answer_cell = ws.cell(row=row, column=answer_col, value=text)
        _copy_cell_style(question_cell, answer_cell)
    return wb


def placements_from_questions_and_answers(
    questions: list[dict],
    question_to_answer: dict[int, str],
) -> list[dict[str, Any]]:
    """Build placement list from questions with source_location and answer map.

    questions: list of {id, source_location, ...}
    question_to_answer: {question_id: answer_text}
    """
    placements = []
    for q in questions:
        qid = q.get("id")
        if qid is None:
            continue
        text = question_to_answer.get(qid, "")
        loc_raw = q.get("source_location")
        if not loc_raw:
            continue
        try:
            loc = json.loads(loc_raw) if isinstance(loc_raw, str) else loc_raw
        except (json.JSONDecodeError, TypeError):
            continue
        placements.append({
            "sheet": loc.get("sheet"),
            "row": loc.get("row", 1),
            "col": loc.get("col", 1),
            "text": text,
        })
    return placements


def create_export_from_questionnaire(
    xlsx_bytes: bytes,
    questions: list[dict],
    question_to_answer: dict[int, str],
    answer_col_offset: int = 1,
) -> bytes:
    """Load original XLSX, place answers, return modified XLSX bytes (EXP-03)."""
    wb = load_workbook_from_bytes(xlsx_bytes)
    placements = placements_from_questions_and_answers(questions, question_to_answer)
    wb = place_answers_into_workbook(wb, placements, answer_col_offset)
    return save_workbook(wb)
